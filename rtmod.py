import openpyxl
import argparse

DEFAULT_2M_OFFSET = '600 khZ'
DEFAULT_70CM_OFFSET = '5.00 MHz'

# insert a column with a default value
def add_filled_column(sheet, col, name, value):
    sheet.insert_cols(idx=col, amount=1)
    sheet.cell(row=1, column=col).value = name
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col, max_col=col):
        row[0].value = value

def fix_ctcss(tone_string):
    "remove Hz from the end and make it an integer"
    return float(tone_string[:-3])

def populate_anytone(workbook, anytone_sheet_name='Anytone', source_sheet_name='Import'):
    "populate the specified sheet in anytone format from a sheet in ft70 format"

    source_sheet = workbook[source_sheet_name]
    anytone_sheet = workbook.create_sheet(anytone_sheet_name)
    
    # populate channel numbers with a counter
    anytone_sheet['A1'].value = 'No.'
    rowctr = 1
    for row in anytone_sheet.iter_rows(min_row=2, max_row=source_sheet.max_row, min_col=1, max_col=1):
        row[0].value = rowctr
        rowctr += 1

    # list of all column headings and defaults, use to set up defaults
    headers =  ["No.","Channel Name","Receive Frequency","Transmit Frequency","Channel Type","Transmit Power","Band Width",
                "CTCSS/DCS Decode","CTCSS/DCS Encode","Contact","Contact Call Type","Contact TG/DMR ID","Radio ID",
                "Busy Lock/TX Permit","Squelch Mode","Optional Signal","DTMF ID","2Tone ID","5Tone ID","PTT ID","Color Code",
                "Slot","Scan List","Receive Group List","PTT Prohibit","Reverse","Simplex TDMA","Slot Suit","AES Digital Encryption",
                "Digital Encryption","Call Confirmation","Talk Around(Simplex)","Work Alone","Custom CTCSS","2TONE Decode","Ranging",
                "Through Mode","APRS RX","Analog APRS PTT Mode","Digital APRS PTT Mode","APRS Report Type",
                "Digital APRS Report Channel","Correct Frequency[Hz]","SMS Confirmation","Exclude channel from roaming","DMR MODE",
                "DataACK Disable","R5toneBot","R5ToneEot","Auto Scan","Ana Aprs Mute","Send Talker Alias"]
    defaults = [1,"KO6DVB",440.65000,445.65000,"A-Analog","Turbo","25K","Off",94.8,"Contact1","Group Call",12345678,"My Radio",
                "Off","Carrier","Off",1,1,1,"Off",1,1,"None","None","Off","Off","Off","Off","Normal Encryption","Off","Off",
                "Off","Off",251.1,0,"Off","On","Off","Off","Off","Off",1,0,"Off",0,0,0,0,0,0,0,0]
    colnum = 2
    for header, default in zip(headers[1:], defaults[1:]):
        add_filled_column(anytone_sheet, colnum, header, default)
        colnum += 1
    # now add name, rx freq, tx freq, channel type, tx pwr, bw, ctcss dec, ctcss enc
    for row in anytone_sheet.iter_rows(min_row=2, max_row=anytone_sheet.max_row, min_col=2, max_col=9):
        row_idx = row[0].row
        for cell in row:
            # handle name
            if cell.column == 2:
                cell.value = source_sheet.cell(row=row_idx, column=8).value
            # handle rx freq
            if cell.column == 3:
                cell.value = source_sheet.cell(row=row_idx, column=2).value
            # handle tx freq
            if cell.column == 4:
                cell.value = source_sheet.cell(row=row_idx, column=3).value
            # skip defaulted 5, 6, 7
            # handle 8 rx ctcss only if col 10 is 'T SQL'
            if cell.column == 8:
                if source_sheet.cell(row=row_idx, column=10).value == 'T SQL':
                    cell.value = fix_ctcss(source_sheet.cell(row=row_idx, column=11).value)
            # handle 9 tx ctcss
            if cell.column == 9:
                cell.value = fix_ctcss(source_sheet.cell(row=row_idx, column=11).value)


def translate_repeaterbook(workbook, sheet_name):
    "translate the named sheet from repeaterbook to ft70 format"
    sheet = workbook[sheet_name]

    # change the offset direction column to two columns: 'Offset Frequency', and 'Offset Direction'
    # 'Offset Direction' is Plus, Minus, Simplex instead of +, -, blank
    if sheet['D1'].value == 'Offset Direction':
        sheet.insert_cols(idx=4, amount=1)
        sheet['D1'].value = 'Offset Frequency'
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=5):
            # transmit_freq, offset_freq, offset_dir
            if row[0].value < 200:
                offset_freq = DEFAULT_2M_OFFSET
            else:
                offset_freq = DEFAULT_70CM_OFFSET
            if row[2].value == '+':
                row[1].value = offset_freq
                row[2].value = 'Plus'
            elif row[2].value == '-':
                row[1].value = offset_freq
                row[2].value = 'Minus'
            else:
                row[2].value = 'Simplex'

        # now add two columns after 'Offset Direction'
        # 'Operating Mode', and AMS
        add_filled_column(sheet, 6, 'Operating Mode', 'FM')
        add_filled_column(sheet, 7, 'AMS', 'On')

        # add the 'Show Name' column and set all the 'On'
        add_filled_column(sheet, 9, 'Show Name', 'On')

        # Change CTCSS to an fp string with an Hz qualifier
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=11, max_col=11):
            tfreq = float(row[0].value)
            tfreq_string = f"{tfreq:.1f} Hz"
            row[0].value = tfreq_string

        # delete the Rx CTCSS column
        sheet.delete_cols(idx=12)

        # delete the RX DCS column
        sheet.delete_cols(idx=13)

        # add remaining defaulted columns before comment field
        add_filled_column(sheet, 13, 'DCS Polarity', 'RN-TN')
        add_filled_column(sheet, 14, 'PR FREQ', '1600 Hz')
        add_filled_column(sheet, 15, 'Tx Power', 'High')
        add_filled_column(sheet, 16, 'Skip', 'Off')
        add_filled_column(sheet, 17, 'Step', 'Auto')
        add_filled_column(sheet, 18, 'Mask', 'Off')
        add_filled_column(sheet, 19, 'Attenuator', 'Off')
        add_filled_column(sheet, 20, 'S-Meter Squelch', 'Off')
        add_filled_column(sheet, 21, 'Bell', 'Off')
        add_filled_column(sheet, 22, 'Half Dev', 'Off')
        add_filled_column(sheet, 23, 'Clock Shift', 'Off')

        # insert the memory bank columns
        for i in range(1, 25):
            add_filled_column(sheet, i + 23, f'BANK {i}', 'Off')

    else:
        print(f"Unexpected value in D1: {sheet['D1'].value}, exiting without modifying")
        exit(1)


def main():
    parser = argparse.ArgumentParser(description='Convert code plugs between formats')
    parser.add_argument('-i', '--input', help='Input file', required=True)
    parser.add_argument('-o', '--output', help='Output file', required=True)
    parser.add_argument('-s', '--sheet', help='sheet to modify or translate', default='Import')
    parser.add_argument('-a', '--anytone', help='sheet name to create in anytone cps format')
    parser.add_argument('-y', '--yaesu', help='sheet in repeaterbook format to modify to RTSystems FT70 format')
    args = parser.parse_args()

    if not args.anytone and not args.yaesu:
        print("Error, must specify either anytone or yaesu")
        exit(1)

    workbook = openpyxl.load_workbook(args.input)

    if args.anytone:
        populate_anytone(workbook, 'Anytone', args.sheet)
    else:
        translate_repeaterbook(workbook, args.sheet)
    
    workbook.save(args.output)

if __name__ == '__main__':
    main()
